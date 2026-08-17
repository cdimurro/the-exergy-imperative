"""Optional native Excel templates, execution, and report workbooks."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping

from .ingestion import _spreadsheet_safe_value, _validate_tabular_headers

_TEMPLATE_DEFINITIONS: dict[str, tuple[tuple[Any, ...], ...]] = {
    "process": (
        ("template", "steam", "", True, "Process template or alias."),
        ("energy", 1000.0, "MWh", False, "Input energy for the reporting period."),
        ("unit", "MWh", "", False, "Energy unit."),
        ("country", "USA", "", False, "Country for grid-factor lookup."),
        ("year", 2025, "", False, "Factor year when available."),
        (
            "improvement_fraction",
            None,
            "fraction",
            False,
            "Optional measured or engineered opportunity fraction.",
        ),
    ),
    "technology-cost": (
        ("name", "industrial heat pump", "", True, "Scenario name."),
        ("capital_cost", 1200000.0, "currency", True, "Installed capital cost."),
        ("annual_output_mwh", 9000.0, "MWh/year", True, "Useful annual output."),
        ("output_name", "useful heat", "", False, "Name of the useful output."),
        ("currency", "USD", "", False, "ISO-style currency label."),
        ("project_life_years", 20, "years", False, "Economic life."),
        ("discount_rate", 0.07, "fraction", False, "Real or nominal discount rate."),
        ("annual_fixed_opex", 30000.0, "currency/year", False, "Fixed operating cost."),
        ("annual_fuel_use_mwh", 3000.0, "MWh/year", False, "Annual purchased energy."),
        ("fuel_price_per_mwh", 70.0, "currency/MWh", False, "First-year fuel price."),
        (
            "carbon_price_per_tonne",
            0.0,
            "currency/tCO2e",
            False,
            "First-year carbon price.",
        ),
        ("annual_emissions_kg_co2e", 0.0, "kg CO2e/year", False, "Annual emissions."),
        (
            "output_value_per_mwh",
            0.0,
            "currency/MWh",
            False,
            "Value or revenue per output.",
        ),
        ("source", "", "", False, "User-owned cost-data source and price basis."),
    ),
    "ghg-boundaries": (
        ("combustion_gases_kg.CO2", 0.0, "kg", False, "Direct combustion CO2."),
        ("process_gases_kg.N2O", 0.0, "kg", False, "Direct process N2O."),
        ("fugitive_gases_kg.CH4-fossil", 0.0, "kg", False, "Fugitive fossil methane."),
        (
            "purchased_energy_co2e_kg",
            0.0,
            "kg CO2e",
            False,
            "Purchased-energy emissions.",
        ),
        (
            "allocated_electricity_heat_co2e_kg",
            None,
            "kg CO2e",
            False,
            "Contextual allocation view; excluded from totals.",
        ),
    ),
    "methane": (
        (
            "annual_methane_mass_kg",
            10000.0,
            "kg/year",
            True,
            "Annual methane available.",
        ),
        (
            "methane_origin",
            "fossil",
            "",
            False,
            "fossil or biogenic accounting basis.",
        ),
        (
            "baseline_mode",
            "vented",
            "",
            False,
            "vented, flared, oxidized, or recovered.",
        ),
        ("project_mode", "recovered", "", False, "Project disposition."),
        (
            "project_efficiency",
            0.90,
            "fraction",
            False,
            "Recovery or destruction efficiency.",
        ),
        (
            "recovered_gas_price_per_mwh",
            35.0,
            "currency/MWh",
            False,
            "Recovered gas value.",
        ),
        ("capital_cost", 75000.0, "currency", False, "Project capital cost."),
        (
            "annual_opex_increase",
            0.0,
            "currency/year",
            False,
            "Incremental annual OPEX.",
        ),
        (
            "carbon_price_per_tonne",
            0.0,
            "currency/tCO2e",
            False,
            "Carbon-value scenario.",
        ),
    ),
    "weather-normalize": (
        (
            "value_field",
            "energy_mwh",
            "",
            True,
            "Column in the Data sheet to normalize.",
        ),
        ("unit", "MWh", "", False, "Metric unit."),
        ("date_field", "date", "", False, "Date column in the Data sheet."),
        (
            "temperature_field",
            "temperature_c",
            "",
            False,
            "Ambient-temperature column.",
        ),
        ("heating_base_c", 18.0, "degC", False, "Heating degree-day base."),
        ("cooling_base_c", 18.0, "degC", False, "Cooling degree-day base."),
        (
            "normal_heating_degree_days",
            0.0,
            "degC-day",
            True,
            "Normal-period HDD total.",
        ),
        (
            "normal_cooling_degree_days",
            0.0,
            "degC-day",
            True,
            "Normal-period CDD total.",
        ),
    ),
    "steam": (
        ("fuel_energy_mwh", 1000.0, "MWh", True, "Fuel energy input."),
        ("boiler_efficiency", 0.80, "fraction", False, "Steam generation efficiency."),
        ("distribution_loss_fraction", 0.05, "fraction", False, "Distribution loss."),
        ("steam_temperature_c", 180.0, "degC", False, "Steam delivery temperature."),
        ("reference_temperature_c", 25.0, "degC", False, "Reference environment."),
    ),
    "heat-pump": (
        ("delivered_heat_mwh", 1000.0, "MWh", True, "Useful delivered heat."),
        ("source_temperature_c", 10.0, "degC", True, "Heat-source temperature."),
        ("sink_temperature_c", 60.0, "degC", True, "Heat delivery temperature."),
        ("cop", 3.0, "", True, "Measured or rated heating COP."),
    ),
    "furnace": (
        ("fuel_energy_mwh", 1000.0, "MWh", True, "Fuel energy input."),
        (
            "thermal_efficiency",
            0.60,
            "fraction",
            False,
            "Useful process-heat efficiency.",
        ),
        ("process_temperature_c", 800.0, "degC", True, "Process temperature."),
        ("exhaust_energy_mwh", 150.0, "MWh", False, "Recoverable exhaust energy."),
        ("exhaust_temperature_c", 300.0, "degC", False, "Exhaust temperature."),
        ("reference_temperature_c", 25.0, "degC", False, "Reference environment."),
    ),
    "refrigeration": (
        ("cooling_delivered_mwh", 1000.0, "MWh", True, "Useful cooling delivered."),
        ("cold_temperature_c", -10.0, "degC", True, "Cold-side temperature."),
        ("ambient_temperature_c", 30.0, "degC", True, "Heat-rejection environment."),
        ("cop", 2.5, "", True, "Measured or rated cooling COP."),
    ),
    "compressed-air": (
        ("electricity_mwh", 1000.0, "MWh", True, "Compressor electricity."),
        (
            "free_air_volume_m3",
            5000000.0,
            "m3",
            True,
            "Free-air volume on the same period.",
        ),
        (
            "delivery_pressure_bar_abs",
            8.0,
            "bar abs",
            True,
            "Header delivery pressure.",
        ),
        (
            "end_use_pressure_bar_abs",
            6.0,
            "bar abs",
            False,
            "Required pressure at use.",
        ),
        ("ambient_pressure_bar_abs", 1.01325, "bar abs", False, "Reference pressure."),
        ("leak_fraction", 0.10, "fraction", False, "Measured or estimated leakage."),
    ),
}


def _openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise ImportError(
            "Native Excel workflows require: pip install exergy-imperative[data]"
        ) from exc
    return openpyxl


def list_excel_templates() -> tuple[str, ...]:
    return tuple(sorted(_TEMPLATE_DEFINITIONS))


@dataclass(frozen=True)
class ExcelTemplatePayload:
    kind: str
    options: Mapping[str, Any]
    data_records: tuple[Mapping[str, Any], ...] = ()
    source_path: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": "1.0",
            "kind": self.kind,
            "options": dict(self.options),
            "data_records": [dict(item) for item in self.data_records],
            "source_path": self.source_path,
        }


def _style_header(cell: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    cell.fill = PatternFill("solid", fgColor="0F766E")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(vertical="center")


def create_excel_template(kind: str, path: str | Path) -> Path:
    """Create an editable native Excel input template for a supported analysis."""

    openpyxl = _openpyxl()
    key = kind.strip().lower().replace("_", "-")
    try:
        definition = _TEMPLATE_DEFINITIONS[key]
    except KeyError as exc:
        raise KeyError(
            f"unknown Excel template {kind!r}; choose from {', '.join(list_excel_templates())}"
        ) from exc
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    instructions = workbook.active
    instructions.title = "README"
    instructions.sheet_view.showGridLines = False
    instructions["A1"] = "Exergy Imperative analysis workbook"
    instructions["A1"].font = openpyxl.styles.Font(size=16, bold=True, color="0F766E")
    instructions["A3"] = "Analysis kind"
    instructions["B3"] = key
    instructions["A5"] = "How to use"
    instructions["A6"] = (
        "1. Edit values in the Inputs sheet; yellow cells are user inputs."
    )
    instructions["A7"] = "2. For weather normalization, add daily observations to Data."
    instructions["A8"] = "3. Run: exergy excel-run INPUT.xlsx --output RESULT.xlsx"
    instructions["A10"] = (
        "The workbook contains no IEA or other licensed publisher data. You are responsible for the rights and terms governing values you enter."
    )
    instructions["A10"].alignment = openpyxl.styles.Alignment(wrap_text=True)
    instructions.column_dimensions["A"].width = 95
    instructions.column_dimensions["B"].width = 28

    inputs = workbook.create_sheet("Inputs")
    inputs.sheet_view.showGridLines = False
    inputs.freeze_panes = "A5"
    inputs["A1"] = "Analysis inputs"
    inputs["A1"].font = openpyxl.styles.Font(size=15, bold=True, color="0F766E")
    inputs.append([])
    inputs.append([])
    inputs.append(["Field", "Value", "Unit", "Required", "Description"])
    for cell in inputs[4]:
        _style_header(cell)
    for field_name, value, unit, required, description in definition:
        inputs.append(
            [field_name, value, unit, "yes" if required else "no", description]
        )
        inputs.cell(inputs.max_row, 2).fill = openpyxl.styles.PatternFill(
            "solid", fgColor="FFF2CC"
        )
        if isinstance(value, float) and 0.0 <= value <= 1.0 and unit == "fraction":
            inputs.cell(inputs.max_row, 2).number_format = "0.0%"
    for column, width in {"A": 38, "B": 22, "C": 20, "D": 12, "E": 70}.items():
        inputs.column_dimensions[column].width = width
    inputs.auto_filter.ref = f"A4:E{inputs.max_row}"

    if key == "weather-normalize":
        data = workbook.create_sheet("Data")
        data.append(["date", "temperature_c", "energy_mwh"])
        for cell in data[1]:
            _style_header(cell)
        data.freeze_panes = "A2"
        data.column_dimensions["A"].width = 18
        data.column_dimensions["B"].width = 20
        data.column_dimensions["C"].width = 18
        data.append([None, None, None])
    workbook.properties.creator = "The Exergy Imperative"
    workbook.properties.title = f"{key} analysis template"
    workbook.save(destination)
    return destination


def _nested_assign(target: dict[str, Any], path: str, value: Any) -> None:
    parts = path.split(".")
    current = target
    for part in parts[:-1]:
        child = current.setdefault(part, {})
        if not isinstance(child, dict):
            raise ValueError(f"input field path conflicts at {part!r}")
        current = child
    current[parts[-1]] = value


def _parse_input_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return None
    if text[0] in "[{" or text in {"true", "false", "null"}:
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            pass
    return text


def read_excel_template(path: str | Path) -> ExcelTemplatePayload:
    """Read a workbook created by :func:`create_excel_template`."""

    openpyxl = _openpyxl()
    source = Path(path)
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        if "README" not in workbook.sheetnames or "Inputs" not in workbook.sheetnames:
            raise ValueError(
                "Excel analysis workbook requires README and Inputs sheets"
            )
        kind = str(workbook["README"]["B3"].value or "").strip().lower()
        if kind not in _TEMPLATE_DEFINITIONS:
            raise ValueError(f"unsupported analysis kind {kind!r}")
        options: dict[str, Any] = {}
        for row_number, row in enumerate(
            workbook["Inputs"].iter_rows(min_row=5, values_only=True), start=5
        ):
            field_name = str(row[0] or "").strip()
            if not field_name:
                continue
            value = _parse_input_value(row[1])
            required = str(row[3] or "").strip().lower() == "yes"
            if value is None:
                if required:
                    raise ValueError(
                        f"required Excel input {field_name!r} is blank at row {row_number}"
                    )
                continue
            _nested_assign(options, field_name, value)
        data_records: list[dict[str, Any]] = []
        if "Data" in workbook.sheetnames:
            rows = workbook["Data"].iter_rows(values_only=True)
            raw_headers = list(next(rows, ()))
            _validate_tabular_headers(raw_headers)
            headers = [str(item).strip() for item in raw_headers]
            for row in rows:
                record = {
                    header: value
                    for header, value in zip(headers, row)
                    if header and value not in {None, ""}
                }
                if record:
                    data_records.append(record)
        return ExcelTemplatePayload(
            kind=kind,
            options=options,
            data_records=tuple(data_records),
            source_path=str(source),
        )
    finally:
        workbook.close()


def run_excel_template(path: str | Path) -> Any:
    """Execute an editable Excel template through the normal public API."""

    payload = read_excel_template(path)
    options = dict(payload.options)
    if payload.kind == "process":
        from .processes import assess_process

        return assess_process(**options)
    if payload.kind == "technology-cost":
        from .economics import evaluate_technology_cost_scenario

        return evaluate_technology_cost_scenario(options)
    if payload.kind == "ghg-boundaries":
        from .ghg import assess_ghg_boundaries

        return assess_ghg_boundaries(**options)
    if payload.kind == "methane":
        from .ghg import assess_methane_project

        return assess_methane_project(**options)
    if payload.kind == "weather-normalize":
        from .weather import normalize_weather_performance

        if not payload.data_records:
            raise ValueError(
                "weather-normalize workbook requires rows in the Data sheet"
            )
        return normalize_weather_performance(payload.data_records, **options)
    from .engineering import (
        analyze_compressed_air,
        analyze_furnace,
        analyze_heat_pump,
        analyze_refrigeration,
        analyze_steam_system,
    )

    dispatch = {
        "steam": analyze_steam_system,
        "heat-pump": analyze_heat_pump,
        "furnace": analyze_furnace,
        "refrigeration": analyze_refrigeration,
        "compressed-air": analyze_compressed_air,
    }
    return dispatch[payload.kind](**options)


def _safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", name).strip()[:31] or "Table"
    candidate = base
    counter = 2
    while candidate.lower() in used:
        suffix = f"-{counter}"
        candidate = base[: 31 - len(suffix)] + suffix
        counter += 1
    used.add(candidate.lower())
    return candidate


def _write_table(sheet: Any, rows: Iterable[Mapping[str, Any]]) -> None:
    materialized = [dict(item) for item in rows]
    columns = list(dict.fromkeys(key for row in materialized for key in row))
    if not columns:
        sheet.append(["No records"])
        return
    sheet.append([_spreadsheet_safe_value(column) for column in columns])
    for cell in sheet[1]:
        _style_header(cell)
    for row in materialized:
        sheet.append([_spreadsheet_safe_value(row.get(column)) for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        maximum = max(
            len(str(column)),
            *(
                len(str(sheet.cell(row, index).value or ""))
                for row in range(2, min(sheet.max_row, 50) + 1)
            ),
        )
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(
            max(maximum + 2, 12), 45
        )


def export_xlsx_report(
    result: Any,
    path: str | Path,
    *,
    title: str | None = None,
) -> Path:
    """Export a supported result to one formatted, native Excel workbook."""

    openpyxl = _openpyxl()
    from openpyxl.chart import BarChart, Reference

    from .reporting import report_view

    view = report_view(result, title=title)
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary["A1"] = _spreadsheet_safe_value(view.title)
    summary["A1"].font = openpyxl.styles.Font(size=16, bold=True, color="0F766E")
    summary["A2"] = _spreadsheet_safe_value(view.subtitle)
    summary["A2"].alignment = openpyxl.styles.Alignment(wrap_text=True)
    summary.append([])
    summary.append(["Metric", "Value", "Unit"])
    for cell in summary[4]:
        _style_header(cell)
    for label, value, unit in view.key_metrics:
        summary.append([_spreadsheet_safe_value(item) for item in (label, value, unit)])
    warning_row = summary.max_row + 2
    summary.cell(warning_row, 1, "Warnings and limitations")
    _style_header(summary.cell(warning_row, 1))
    for warning in view.warnings:
        summary.append([_spreadsheet_safe_value(warning)])
        summary.cell(summary.max_row, 1).alignment = openpyxl.styles.Alignment(
            wrap_text=True
        )
    source_row = summary.max_row + 2
    summary.cell(source_row, 1, "Sources")
    _style_header(summary.cell(source_row, 1))
    summary.append(["Source ID", "Title", "URL"])
    for item in view.sources:
        summary.append(
            [
                _spreadsheet_safe_value(item.get(key, ""))
                for key in ("source_id", "title", "url")
            ]
        )
    summary.column_dimensions["A"].width = 52
    summary.column_dimensions["B"].width = 24
    summary.column_dimensions["C"].width = 22
    summary.freeze_panes = "A5"

    used = {"summary"}
    for table_name, rows in view.tables.items():
        sheet = workbook.create_sheet(_safe_sheet_name(table_name, used))
        sheet.sheet_view.showGridLines = False
        _write_table(sheet, rows)

    chart_data = workbook.create_sheet("Chart Data")
    chart_data.sheet_state = "hidden"
    chart_column = 1
    for index, chart_spec in enumerate(view.charts, start=1):
        chart_data.cell(1, chart_column, "Label")
        chart_data.cell(1, chart_column + 1, _spreadsheet_safe_value(chart_spec.title))
        for row_index, (label, value) in enumerate(chart_spec.values, start=2):
            chart_data.cell(row_index, chart_column, _spreadsheet_safe_value(label))
            chart_data.cell(row_index, chart_column + 1, value)
        if chart_spec.values:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
            chart.title = chart_spec.title
            chart.y_axis.title = ""
            chart.x_axis.title = chart_spec.unit
            data = Reference(
                chart_data,
                min_col=chart_column + 1,
                min_row=1,
                max_row=1 + len(chart_spec.values),
            )
            categories = Reference(
                chart_data,
                min_col=chart_column,
                min_row=2,
                max_row=1 + len(chart_spec.values),
            )
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 7
            chart.width = 12
            summary.add_chart(chart, f"E{2 + (index - 1) * 15}")
        chart_column += 3

    metadata = workbook.create_sheet("Metadata")
    metadata.sheet_view.showGridLines = False
    metadata.append(["Field", "Value"])
    for cell in metadata[1]:
        _style_header(cell)
    metadata.append(["schema_version", "1.0"])
    metadata.append(["title", _spreadsheet_safe_value(view.title)])
    metadata.append(["subtitle", _spreadsheet_safe_value(view.subtitle)])
    metadata.append(["warnings", json.dumps(view.warnings, ensure_ascii=False)])
    payload_text = json.dumps(view.payload, ensure_ascii=False, sort_keys=True)
    for index in range(0, len(payload_text), 30000):
        metadata.append(
            [
                f"source_payload_{index // 30000 + 1}",
                _spreadsheet_safe_value(payload_text[index : index + 30000]),
            ]
        )
    metadata.column_dimensions["A"].width = 28
    metadata.column_dimensions["B"].width = 100
    for cell in metadata["B"]:
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    workbook.properties.creator = "The Exergy Imperative"
    workbook.properties.title = view.title
    workbook.save(destination)
    return destination


def export_xlsx_ingestion(
    result: Any,
    path: str | Path,
    *,
    metadata: Mapping[str, Any] | None = None,
    source_records: Any | None = None,
) -> Path:
    """Write raw, normalized, issue, and mapping sheets for an ingestion result."""

    openpyxl = _openpyxl()
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    tables = []
    if source_records is not None:
        tables.append(("Source Data", source_records))
    tables.extend(
        (
            ("Raw Data", result.raw_records),
            ("Normalized Data", result.records),
            ("Data Quality Issues", (item.to_dict() for item in result.issues)),
        )
    )
    for name, rows in tables:
        sheet = workbook.create_sheet(name)
        sheet.sheet_view.showGridLines = False
        _write_table(sheet, rows)
    mapping = workbook.create_sheet("Mapping")
    mapping.sheet_view.showGridLines = False
    _write_table(mapping, (item.to_dict() for item in result.mapping.fields))
    mapping_plan = workbook.create_sheet("Mapping Plan")
    mapping_plan.sheet_view.showGridLines = False
    plan_payload = result.mapping.to_dict()
    plan_payload["missing_policy"] = result.missing_policy
    _write_table(
        mapping_plan,
        ({"setting": name, "value": value} for name, value in plan_payload.items()),
    )
    if metadata:
        provenance = workbook.create_sheet("Provenance")
        provenance.sheet_view.showGridLines = False
        _write_table(
            provenance,
            (
                {
                    "field": name,
                    "value": (
                        json.dumps(value, ensure_ascii=False)
                        if isinstance(value, (dict, list, tuple))
                        else value
                    ),
                }
                for name, value in metadata.items()
            ),
        )
    workbook.save(destination)
    return destination
